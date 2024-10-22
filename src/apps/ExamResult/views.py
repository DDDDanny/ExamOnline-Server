# -*- coding: utf-8 -*-
# @Time    : 2024/02/24 20:05:21
# @Author  : DannyDong
# @File    : views.py
# @Describe: ExamResult应用视图层

import os

from openpyxl import Workbook
from openpyxl.styles import Font
from django.db.models import Sum
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination

from .models import ExamResult, ExamResultDetail
from src.apps.ExamManagement.models import Exam
from src.apps.ExamManagement.serializers import ExamSerializer
from src.apps.QuestionManagement.models import Questions, ErrorArchive
from src.apps.PaperManagement.models import PaperQuestions
from src.apps.PaperManagement.serializers import PaperQuestionsSerializer
from .serializers import ExamResultSerializer
from .serializers import ExamResultDetailSerializer
from src.utils.response_utils import ResponseCode, api_response


class ExamResultBaseView(APIView):
    # JWT校验
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """post 批量创建考试结果信息
        Args:
            request (Object): { "exam_id": 考试ID, "student_ids": 学生IDs }
        """
        student_ids = request.data['student_ids']
        exam_result_list = []
        if len(student_ids) == 0:
            return api_response(ResponseCode.SUCCESS, '考试关联的学生为空！')
        else:
            for s_id in student_ids:
                exam_result_list.append({
                    'exam_id': request.data['exam_id'],
                    'student_id': s_id
                })
            serializer = ExamResultSerializer(data=exam_result_list, many=True)
            if serializer.is_valid():
                # 如果新数据校验通过，对老数据进行删除操作
                old_data = ExamResult.objects.filter(exam_id=request.data['exam_id'])
                old_data_length = len(old_data)
                deleted_count, _ = old_data.delete()
                if deleted_count != old_data_length:
                    return api_response(ResponseCode.BAD_REQUEST, '创建失败！数据处理失败！', serializer.errors)
                # 老数据删除后，新数据存储
                serializer.save()
                data = Response(serializer.data)
                return api_response(ResponseCode.SUCCESS, '创建成功', data.data)
            else:
                return api_response(ResponseCode.BAD_REQUEST, '创建失败', serializer.errors)

    def delete(self, _, **kwargs):
        """delete 删除考试结果信息（物理删除）
        Args:
            _ (-): 缺省参数
            id (Object): 考试ID
        """
        try:
            exam_instance = ExamResult.objects.get(id=kwargs['id'])
        except ExamResult.DoesNotExist:
            return api_response(ResponseCode.NOT_FOUND, '考试结果不存在，删除失败！')
        if exam_instance.start_time is not None:
            return api_response(ResponseCode.BAD_REQUEST, '该学生已经参加考试，无法删除！')
        else:
            exam_instance.delete()
            # 返回成功响应
            return api_response(ResponseCode.SUCCESS, '删除成功！')

    def put(self, request, **kwargs):
        """put 编辑考试结果信息
        Args:
            id (str): 考试结果ID
            request (Object): 请求参数
        """
        try:
            # 获取需要编辑的考试结果实例
            exam_result_instance = ExamResult.objects.get(id=kwargs['id'])
        except ExamResult.DoesNotExist:
            return api_response(ResponseCode.NOT_FOUND, '编辑失败！考试结果不存在，无法进行编辑！')
        serializer = ExamResultSerializer(exam_result_instance, request.data, partial=True)
        # 检查更新后的数据是否符合规则校验
        if serializer.is_valid():
            # 保存验证过的数据以更新现有的 ExamResult 实例
            serializer.save()
            # 返回成功响应，包含序列化后的数据和 HTTP 200 OK 状态
            data = Response(serializer.data)
            return api_response(ResponseCode.SUCCESS, '编辑成功', data.data)
        else:
            # 返回错误响应，包含验证错误和 HTTP 400 Bad Request 状态
            return api_response(ResponseCode.BAD_REQUEST, '编辑失败！存在校验失败的字段', serializer.error_messages)

    def get(self, request, **kwargs):
        """get 查询考试结果列表信息（根据考试结果ID获取考试结果详情）
        Args:
            request (Object): 请求参数
            kwargs[id] (str): 考试结果ID
        """
        if len(kwargs.items()) != 0:
            try:
                # 获取指定考试结果实例
                exam_result_instance = ExamResult.objects.get(id=kwargs['id'])
            except ExamResult.DoesNotExist:
                # 考试结果不存在，返回错误响应和 HTTP 404 Not Found 状态
                return api_response(ResponseCode.NOT_FOUND, '考试结果不存在！')
            # 序列化试题详情数据
            serializer = ExamResultSerializer(exam_result_instance)
            # 返回序列化后的试题详情数据
            data = Response(serializer.data)
            return api_response(ResponseCode.SUCCESS, '查询考试结果成功', data.data)
        else:
            # 定义查询参数和它们对应的模型字段
            query_params_mapping = {
                'exam_id': 'exam_id',
                'result_mark': 'result_mark'
                # 添加其他查询参数和字段的映射
            }
            # 构建查询条件的字典
            filters = {}
            for param, field in query_params_mapping.items():
                value = request.query_params.get(param, None)
                if value is not None and value != '':
                    filters[field] = value
            # 执行查询
            queryset = ExamResult.objects.filter(**filters).order_by('-result_mark')
            # 序列化试题数据
            serializer = ExamResultSerializer(queryset, many=True)
            serializer_data = serializer.data
            # 获取查询参数
            filter_stu_id = request.query_params.get('student_id', None)
            filter_stu_name = request.query_params.get('name', None)
            # 二次筛选
            if filter_stu_id is not None or filter_stu_name is not None:
                filter_stu_id = filter_stu_id.strip() if filter_stu_id else None
                filter_stu_name = filter_stu_name.strip() if filter_stu_name else None

                def filter_item(item):
                    if filter_stu_id and filter_stu_id not in item['student_info']['student_id']:
                        return False
                    if filter_stu_name and filter_stu_name not in item['student_info']['name']:
                        return False
                    return True

                serializer_data = [item for item in serializer_data if filter_item(item)]
            # 返回序列化后的数据
            data = Response(serializer_data)
            resp = { 'total': len(data.data), 'data': data.data }
            return api_response(ResponseCode.SUCCESS, '查询成功', resp)


class ExamResultStudentView(APIView):
    # JWT校验
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """get 查询考试结果（学生端）
        Args:
            request (Object): 请求参数
        """
        student_id = request.query_params.get("student_id", None)
        queryset = ExamResult.objects.filter(student_id=student_id).order_by('-updated_at')
        # 序列化数据
        serializer_data = ExamResultSerializer(queryset, many=True).data
        title_filter = request.query_params.get("title", None)  # 获取标题筛选参数
        filter_datas = []
        # 根据 exam_info.title 进行筛选
        if title_filter:
            for item in serializer_data:
                if title_filter in item['exam_info']['title']:
                    filter_datas.append(item)
        else:
            filter_datas = serializer_data

        # 实例化分页器并配置参数
        paginator = PageNumberPagination()
        paginator.page_size = int(request.query_params.get('pageSize', 50))
        paginator.page_query_param = 'currentPage'
        # 进行分页处理
        paginated_queryset = paginator.paginate_queryset(filter_datas, request)
        resp = { 'total': len(filter_datas), 'data': paginated_queryset }
        return api_response(ResponseCode.SUCCESS, '查询成功', resp)


class ExamOnlineGetResultView(APIView):
    
    def get(self, request):
        """get 在线考试页面使用 - 获取考生考试信息
        Args:
            request (Object): 
                examId: 考试ID
                studentId: 学生ID
        """
        exam_id = request.query_params.get('examId')
        student_id = request.query_params.get('studentId')
        queryset = ExamResult.objects.filter(exam_id=exam_id).filter(student_id=student_id)
        if len(queryset) <= 0:
            return api_response(ResponseCode.BAD_REQUEST, '没有查询到对应的考试信息，请刷新后重试！')
        # 序列化考试数据
        serializer = ExamResultSerializer(queryset.first())
        return api_response(ResponseCode.SUCCESS, '查询成功', serializer.data)
        

class ExamResultDetailBaseView(APIView):
    # JWT校验
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """post 创建考试结果详情信息
        Args:
            request (Object): 请求参数
        """
        meta_data = request.data
        # 获取考试ID、试卷ID
        exam_id = ExamResult.objects.filter(id=meta_data['exam_result_id']).first().exam_id
        exam_info = Exam.objects.filter(id=exam_id).first()
        # 及格分数
        pass_mark = exam_info.pass_mark
        # 试卷总分
        sum_marks = PaperQuestions.objects.filter(
                    paper_id=exam_info.paper_id).aggregate(actual_total=Sum('marks'))
        paper_questions = PaperQuestions.objects.filter(paper_id=exam_info.paper_id)
        serializer_paper_questions = PaperQuestionsSerializer(paper_questions, many=True).data
        answers = meta_data['answers']
        result_detail_list = []
        result_total_mark = 0
        for key in answers.keys():
            result_record = { 'exam_result_id': meta_data['exam_result_id'], 'question_id': key, 'solution': answers[key], 'mark': 0 }
            # 若学生没有答题，直接0分
            if answers[key] is not None:
                filter_res = list(filter(lambda x: x['question_id'] == key, serializer_paper_questions))
                # 判断答案是否正确
                if filter_res[0]['question_detail']['answer'] == answers[key]:
                    result_record['mark'] = filter_res[0]['marks']
            result_total_mark += result_record['mark']
            result_detail_list.append(result_record)
        serializer = ExamResultDetailSerializer(data=result_detail_list, many=True)
        if serializer.is_valid():
            serializer.save()
            data = Response(serializer.data)
            return api_response(ResponseCode.SUCCESS, '创建成功', {
                "data": data.data, 
                "result_total_mark": result_total_mark, 
                "sum_marks": sum_marks['actual_total'],
                "percentage": (result_total_mark / sum_marks['actual_total']) * 100,
                "pass_mark": pass_mark
            })
        else:
            return api_response(ResponseCode.BAD_REQUEST, '创建失败', serializer.errors)

    def get(self, request, **kwargs):
        """get 根据考试结果ID获取详情信息
        Args:
            _ (——): 缺省参数
            id：试卷ID
        """
        student_id = request.query_params.get('student_id', None)
        # 获取考试结果ID为入参的详情实例
        exam_result_detal_instance = ExamResultDetail.objects.filter(exam_result_id=kwargs['id'])
        if len(exam_result_detal_instance) == 0:
            return api_response(ResponseCode.NOT_FOUND, '没有找到该考试结果的 详情信息！')
        else:
            serializer = ExamResultDetailSerializer(exam_result_detal_instance, many=True)
            for item in serializer.data:
                # 若存在学生ID，判断是否收藏该题目
                if student_id:
                    collected = ErrorArchive.objects.filter(collector=student_id).filter(question_id=item['question_id']).values()
                    item['is_error_archive'] = True if len(collected) != 0 else False
                answer = Questions.objects.filter(id=item['question_id']).first().answer
                item['is_true'] = True if answer == item['solution'] else False
                item['reference_answer'] = answer
            data = Response(serializer.data)
            return api_response(ResponseCode.SUCCESS, '获取考试结果详情成功', data.data)


class GenerateExamResultExcel(APIView):
    
    def post(self, request):
        try:
            exam_id = request.data['exam_id']
            exam_instance = Exam.objects.filter(id=exam_id).first()
            exam_serializer = ExamSerializer(exam_instance)
            exam_result_instance = ExamResult.objects.filter(exam_id=exam_id).order_by('-result_mark')
            exam_result_datas = ExamResultSerializer(exam_result_instance, many=True).data
            
            # 创建一个新的工作簿
            wb = Workbook()
            ws = wb.active

            # 定义总览标题和对应的数据
            headers = [['考试名称', '试卷名称', '总分'], ['考试开始时间', '考试结束时间', '考试人数']]
            datas = [[
                exam_serializer.data['title'],
                exam_serializer.data['paper_info']['title'],
                exam_serializer.data['paper_info']['total_marks']
            ],[
                exam_serializer.data['start_time'],
                exam_serializer.data['end_time'],
                len(exam_result_datas)
            ]]
            # 总览数据填充
            for index in range(len(headers)):
                # 写入标题并设置加粗
                for col, (header, data) in enumerate(zip(headers[index], datas[index]), start=1):
                    if col == 1:
                        cell = ws.cell(row=index + 1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell = ws.cell(row=index + 1, column=col + 1, value=data)
                    else:
                        cell = ws.cell(row=index + 1, column=2 * col - 1, value=header)
                        cell.font = Font(bold=True)
                        cell = ws.cell(row=index + 1, column=2 * col, value=data)
            
            # 定义主体表头数据并填充
            students_result_header = ['学号', '学生姓名', '得分', '考试状态', '考试开始时间', '考试结束时间']
            for index, item in enumerate(students_result_header):
                cell = ws.cell(row=4, column=index+1, value=item)
                cell.font = Font(bold=True)
            
            # 主体数据填充
            for index, item in enumerate(exam_result_datas):
                cell = ws.cell(row=index+5, column=1, value=item['student_info']['student_id'])
                cell = ws.cell(row=index+5, column=2, value=item['student_info']['name'])
                cell = ws.cell(row=index+5, column=3, value=item['result_mark'])
                cell = ws.cell(row=index+5, column=4, value='正常' if item['ending_status'] else '异常退出' )
                cell = ws.cell(row=index+5, column=5, value=item['start_time'])
                cell = ws.cell(row=index+5, column=6, value=item['end_time'])

            # 设置所有行的高度为 30
            for row in range(1, len(exam_result_datas) + 6):
                ws.row_dimensions[row].height = 30
                
            # 设置列 A 到 F 的宽度
            column_width = 35  # 设置的宽度值
            for col in range(ord('A'), ord('F') + 1):  # A到F的ASCII码范围
                ws.column_dimensions[chr(col)].width = column_width

            # 指定目录路径
            directory = 'ExamResultFiles/'
            os.makedirs(directory, exist_ok=True)  # 确保目录存在
            file_name = '{}.xlsx'.format(exam_id)
            # 保存文件
            wb.save(os.path.join(directory, file_name))
            return api_response(ResponseCode.SUCCESS, '生成成功', file_name)
        except Exception as e:
            return api_response(ResponseCode.BAD_REQUEST, '生成失败!', e)
        

if __name__ == '__main__':
    pass
